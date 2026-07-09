from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment

from backend.report_utils import write_report_sheet


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    """
    Convert DataFrame into Excel bytes.
    Newlines inside cells are preserved.
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="extracted_data")

        worksheet = writer.sheets["extracted_data"]

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = 28

    output.seek(0)
    return output.getvalue()


def dataframe_to_excel_bytes_with_report(
    df: pd.DataFrame,
    report: dict[str, Any],
) -> bytes:
    """
    Convert DataFrame + extraction report into a two-sheet Excel workbook:

        Sheet 1: "Extraction Report"  (readable summary, see report_utils)
        Sheet 2: "Extracted Table"    (the output table, same formatting
                                        as dataframe_to_excel_bytes)

    Sheet order is guaranteed: report first, table second.
    """

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Report sheet must be created first so it ends up first in the
        # workbook (openpyxl appends new sheets at the end by default).
        write_report_sheet(writer, report, sheet_name="Extraction Report")

        df.to_excel(writer, index=False, sheet_name="Extracted Table")

        worksheet = writer.sheets["Extracted Table"]

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            worksheet.column_dimensions[column_letter].width = 28

        # pandas' ExcelWriter (openpyxl engine) starts with one default
        # empty sheet named "Sheet". Remove it if it's still untouched,
        # so the workbook only contains the two intended sheets in order.
        workbook = writer.book

        if "Sheet" in workbook.sheetnames:
            default_sheet = workbook["Sheet"]

            if (
                default_sheet.max_row == 1
                and default_sheet.max_column == 1
                and default_sheet["A1"].value is None
            ):
                del workbook["Sheet"]

    output.seek(0)
    return output.getvalue()