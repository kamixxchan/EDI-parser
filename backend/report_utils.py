"""
Extraction Report builder.

Turns a completed extraction (the result DataFrame plus a few run
details) into a structured report dict that both the Streamlit UI and
the Excel export can render. Pure Python / pandas only — no Streamlit
dependency here, so it stays testable and reusable outside the app,
matching the other backend/ modules.

Every function here is defensive: a missing column, an empty table, or
a missing run detail (filename, page range, processing time) degrades
to a safe fallback value instead of raising.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any

import pandas as pd


UNKNOWN = "Unknown"
NOT_AVAILABLE = "N/A"

# How many distinct review reasons to surface in the report.
TOP_REVIEW_REASONS_LIMIT = 10


def _safe_str(value: Any, fallback: str = UNKNOWN) -> str:
    text = str(value).strip() if value is not None else ""
    return text if text else fallback


def _format_page_range(page_start: int | None, page_end: int | None) -> str:
    if page_start is None or page_end is None:
        return NOT_AVAILABLE

    try:
        start = int(page_start)
        end = int(page_end)
    except (TypeError, ValueError):
        return NOT_AVAILABLE

    if end < start:
        return NOT_AVAILABLE

    return f"{start}\u2013{end}" if end != start else str(start)


def _format_processing_time(seconds: float | None) -> str:
    if seconds is None:
        return NOT_AVAILABLE

    try:
        seconds = float(seconds)
    except (TypeError, ValueError):
        return NOT_AVAILABLE

    if seconds < 0:
        return NOT_AVAILABLE

    return f"{seconds:.1f} seconds"


def _count_pages_processed(page_start: int | None, page_end: int | None) -> int | str:
    if page_start is None or page_end is None:
        return NOT_AVAILABLE

    try:
        start = int(page_start)
        end = int(page_end)
    except (TypeError, ValueError):
        return NOT_AVAILABLE

    if end < start:
        return NOT_AVAILABLE

    return end - start + 1


def _build_file_information(
    pdf_filename: str | None,
    extraction_mode: str | None,
    page_start: int | None,
    page_end: int | None,
    generated_at: datetime | None,
) -> dict[str, str]:
    generated_at = generated_at or datetime.now()

    return {
        "PDF File Name": _safe_str(pdf_filename),
        "Extraction Mode": _safe_str(extraction_mode),
        "Page Range": _format_page_range(page_start, page_end),
        "Generated At": generated_at.strftime("%Y-%m-%d %H:%M:%S"),
    }


def _build_processing_summary(
    df: pd.DataFrame | None,
    page_start: int | None,
    page_end: int | None,
    processing_time_seconds: float | None,
) -> dict[str, Any]:
    total_rows = 0 if df is None else len(df)

    if df is not None and "segment_id" in df.columns:
        segment_values = df["segment_id"].astype(str).str.strip()
        segment_values = segment_values[segment_values != ""]
        total_segments = int(segment_values.nunique())
    else:
        total_segments = 0

    return {
        "Total Processing Time": _format_processing_time(processing_time_seconds),
        "Total Pages Processed": _count_pages_processed(page_start, page_end),
        "Total Rows Extracted": total_rows,
        "Total Segments Detected": total_segments,
    }


def _build_segment_summary(df: pd.DataFrame | None) -> pd.DataFrame:
    columns = ["Segment ID", "Segment Name", "Rows Extracted"]

    if df is None or df.empty or "segment_id" not in df.columns:
        return pd.DataFrame(columns=columns)

    working = df.copy()
    working["segment_id"] = (
        working["segment_id"].astype(str).str.strip().replace("", UNKNOWN)
    )

    if "segment_name" in working.columns:
        working["segment_name"] = (
            working["segment_name"].astype(str).str.strip().replace("", UNKNOWN)
        )
    else:
        working["segment_name"] = UNKNOWN

    grouped = (
        working.groupby("segment_id", dropna=False, sort=False)
        .agg(
            **{
                "Segment Name": ("segment_name", "first"),
                "Rows Extracted": ("segment_id", "count"),
            }
        )
        .reset_index()
        .rename(columns={"segment_id": "Segment ID"})
        .reset_index(drop=True)
    )

    return grouped[columns]


def _build_quality_summary(df: pd.DataFrame | None) -> dict[str, Any]:
    total_rows = 0 if df is None else len(df)

    if df is not None and "needs_review" in df.columns:
        review_count = int(df["needs_review"].fillna(False).astype(bool).sum())
    else:
        review_count = 0

    review_percentage = (review_count / total_rows * 100) if total_rows else 0.0

    return {
        "total_rows": total_rows,
        "Rows Requiring Review": review_count,
        "Review Percentage": f"{review_percentage:.2f}%",
        "_review_percentage_value": review_percentage,
    }


def _build_status(total_rows: int, review_percentage: float) -> dict[str, str]:
    if total_rows == 0:
        status = "No Data Extracted"
        explanation = (
            "No rows were extracted from this document. Check the page range, "
            "source columns, and extraction mode before trying again."
        )
    elif review_percentage < 5:
        status = "Excellent"
        explanation = (
            "The extraction result looks highly reliable. A final skim of any "
            "flagged rows is still recommended before export."
        )
    elif review_percentage < 15:
        status = "Good"
        explanation = (
            "The extraction result looks mostly reliable. Please review all "
            "rows marked as needing review before using the final Excel output."
        )
    elif review_percentage < 30:
        status = "Moderate"
        explanation = (
            "A meaningful portion of rows need review. Check flagged rows "
            "carefully, and consider whether the column configuration matches "
            "this document before exporting."
        )
    else:
        status = "Needs Careful Review"
        explanation = (
            "A large share of rows need review. Treat this output as a draft: "
            "verify the page range, source columns, and key/multi-line column "
            "settings, then re-run extraction if needed before relying on it."
        )

    return {"status": status, "explanation": explanation}


def _build_review_reason_summary(df: pd.DataFrame | None) -> pd.DataFrame:
    columns = ["Review Reason", "Count"]

    if df is None or df.empty or "review_reason" not in df.columns:
        return pd.DataFrame(columns=columns)

    reasons = df["review_reason"].astype(str).str.strip()
    reasons = reasons[reasons != ""]

    if reasons.empty:
        return pd.DataFrame(columns=columns)

    counts = (
        reasons.value_counts()
        .head(TOP_REVIEW_REASONS_LIMIT)
        .rename_axis("Review Reason")
        .reset_index(name="Count")
    )

    return counts[columns]


def _build_warnings(
    df: pd.DataFrame | None,
    quality_summary: dict[str, Any],
    review_reason_summary: pd.DataFrame,
) -> list[str]:
    warnings: list[str] = []

    review_count = quality_summary.get("Rows Requiring Review", 0)

    if review_count:
        warnings.append(
            f"Review all {review_count} row(s) where needs_review is set "
            "before downloading the final file."
        )

    warnings.append(
        "Check rows with page break issues or incomplete qualifier information."
    )
    warnings.append(
        "Validate segment IDs and segment names if the document has unusual "
        "formatting."
    )
    warnings.append(
        "Confirm that the final output table matches the expected mapping "
        "spec structure before submitting or sharing."
    )

    if df is not None and not df.empty and "segment_id" not in df.columns:
        warnings.append(
            "This extraction mode does not report segment IDs, so the "
            "Segment Summary section is unavailable for this result."
        )

    if not review_reason_summary.empty:
        top_reason = review_reason_summary.iloc[0]
        warnings.append(
            f"Most common review reason: \"{top_reason['Review Reason']}\" "
            f"({int(top_reason['Count'])} row(s))."
        )

    return warnings


def build_extraction_report(
    df: pd.DataFrame | None,
    pdf_filename: str | None,
    extraction_mode: str | None,
    page_start: int | None,
    page_end: int | None,
    processing_time_seconds: float | None,
    generated_at: datetime | None = None,
) -> dict[str, Any]:
    """
    Build a structured extraction report from the result DataFrame and
    the run's metadata.

    Safe for every edge case: df is None/empty, needs_review/
    review_reason/segment_id/segment_name columns are missing, and
    pdf_filename/page_start/page_end/processing_time_seconds are None.

    Returns a dict with:
        file_info               dict[str, str]
        processing_summary      dict[str, Any]
        segment_summary         pd.DataFrame (Segment ID/Name/Rows Extracted)
        quality_summary         dict[str, Any]
        status                  dict[str, str] ("status", "explanation")
        warnings                list[str]
        review_reason_summary   pd.DataFrame (Review Reason/Count)
    """

    if df is not None and not isinstance(df, pd.DataFrame):
        df = None

    file_info = _build_file_information(
        pdf_filename=pdf_filename,
        extraction_mode=extraction_mode,
        page_start=page_start,
        page_end=page_end,
        generated_at=generated_at,
    )

    processing_summary = _build_processing_summary(
        df=df,
        page_start=page_start,
        page_end=page_end,
        processing_time_seconds=processing_time_seconds,
    )

    segment_summary = _build_segment_summary(df)
    quality_summary = _build_quality_summary(df)
    review_reason_summary = _build_review_reason_summary(df)

    status = _build_status(
        total_rows=quality_summary["total_rows"],
        review_percentage=quality_summary["_review_percentage_value"],
    )

    warnings = _build_warnings(
        df=df,
        quality_summary=quality_summary,
        review_reason_summary=review_reason_summary,
    )

    quality_summary_display = {
        k: v for k, v in quality_summary.items() if not k.startswith("_") and k != "total_rows"
    }

    return {
        "file_info": file_info,
        "processing_summary": processing_summary,
        "segment_summary": segment_summary,
        "quality_summary": quality_summary_display,
        "status": status,
        "warnings": warnings,
        "review_reason_summary": review_reason_summary,
    }


def write_report_sheet(
    writer: "pd.ExcelWriter",
    report: dict[str, Any],
    sheet_name: str = "Extraction Report",
) -> Any:
    """
    Write the extraction report into its own worksheet on an already-open
    pandas ExcelWriter (openpyxl engine), formatted as readable sections
    rather than a raw DataFrame dump.

    Returns the created worksheet.
    """

    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = writer.book
    worksheet = workbook.create_sheet(title=sheet_name)

    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    row = 1

    def write_title(text: str) -> None:
        nonlocal row
        cell = worksheet.cell(row=row, column=1, value=text)
        cell.font = title_font
        row += 2

    def write_section(text: str) -> None:
        nonlocal row
        cell = worksheet.cell(row=row, column=1, value=text)
        cell.font = section_font
        row += 1

    def write_kv_table(data: dict[str, Any]) -> None:
        nonlocal row
        for key, value in data.items():
            key_cell = worksheet.cell(row=row, column=1, value=str(key))
            key_cell.font = header_font
            value_cell = worksheet.cell(row=row, column=2, value=str(value))
            value_cell.alignment = wrap
            row += 1
        row += 1

    def write_table(df: pd.DataFrame, empty_message: str | None = None) -> None:
        nonlocal row

        if df is None or df.empty:
            if empty_message:
                worksheet.cell(row=row, column=1, value=empty_message)
                row += 2
            return

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=row, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill

        row += 1

        for _, record in df.iterrows():
            for col_idx, col_name in enumerate(df.columns, start=1):
                value = record[col_name]
                cell = worksheet.cell(
                    row=row,
                    column=col_idx,
                    value=None if pd.isna(value) else value,
                )
                cell.alignment = wrap
            row += 1

        row += 1

    write_title("Extraction Report")

    write_section("A. File Information")
    write_kv_table(report["file_info"])

    write_section("B. Processing Summary")
    write_kv_table(report["processing_summary"])

    write_section("C. Segment Summary")
    write_table(
        report["segment_summary"],
        empty_message="No segment information available for this extraction mode.",
    )

    write_section("D. Quality Summary")
    write_kv_table(report["quality_summary"])

    write_section("E. Extraction Confidence / Status")
    status_cell = worksheet.cell(row=row, column=1, value="Status")
    status_cell.font = header_font
    worksheet.cell(row=row, column=2, value=report["status"]["status"])
    row += 1
    explanation_cell = worksheet.cell(row=row, column=1, value=report["status"]["explanation"])
    explanation_cell.alignment = wrap
    row += 2

    write_section("F. Warnings and Recommendations")

    for warning_text in report["warnings"]:
        cell = worksheet.cell(row=row, column=1, value=f"\u2022 {warning_text}")
        cell.alignment = wrap
        row += 1

    row += 1

    if not report["review_reason_summary"].empty:
        write_section("Most Common Review Reasons")
        write_table(report["review_reason_summary"])

    worksheet.column_dimensions["A"].width = 55
    worksheet.column_dimensions["B"].width = 45
    worksheet.column_dimensions["C"].width = 20
    worksheet.column_dimensions["D"].width = 20

    return worksheet